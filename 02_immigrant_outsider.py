"""
02_immigrant_outsider.py
Reproduces every derived number and the workbook for the immigrant/outsider theme.

The presence/type/function judgements are RESEARCHER-CODED (close reading of all 61
interviews) and live in inputs/immigrant_outsider_coded.csv — that file is the data.
This script (a) recomputes all summary statistics from it, (b) runs an independent
discourse-marker cross-check on the raw passages, and (c) writes the .xlsx workbook.

Run:  python 02_immigrant_outsider.py
Out:  outputs/immigrant_outsider_summary.csv, outputs/immigrant_outsider_coded.xlsx
"""
import re, pandas as pd
from common import load_evidence, inp, out
from markers import THEME_PATTERNS

def summarise(df):
    n = len(df)
    present = df[df.outsider_present == "yes"]
    mig = df[df.literal_migration == "yes"]
    print(f"N interviews: {n}")
    print(f"outsider_present = yes: {len(present)} ({len(present)/n:.0%})")
    print(f"literal_migration = yes: {len(mig)} ({len(mig)/n:.0%})")
    for g in ("F", "M"):
        sub = df[df.gender == g]; p = sub[sub.outsider_present == "yes"]
        print(f"  {g}: present {len(p)}/{len(sub)} ({len(p)/len(sub):.0%}); "
              f"migration {sum(sub.literal_migration=='yes')}/{len(sub)}")
    print("\nprimary_type (present):", present.primary_type.value_counts().to_dict())
    print("function (present):", present.function.value_counts().to_dict())
    print("function x gender (present):")
    for g in ("F", "M"):
        print(f"  {g}:", present[present.gender == g].function.value_counts().to_dict())

def marker_crosscheck():
    """Independent keyword estimate of literal-migration prevalence, to validate the coding."""
    _, epi = load_evidence()
    pat = THEME_PATTERNS["immigrant/outsider"]
    hit = epi["t"].apply(lambda s: 1 if re.search(pat, s) else 0)
    print(f"\n[cross-check] discourse-marker migration/outsider prevalence: {hit.mean():.0%} "
          f"(compare to hand-coded literal_migration above)")

def build_xlsx(df):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    AR = "Arial"; hf = PatternFill("solid", fgColor="1F3864"); hfont = Font(name=AR, bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="BFBFBF"); bd = Border(thin, thin, thin, thin)
    wrap = Alignment(wrap_text=True, vertical="top"); ctr = Alignment(horizontal="center", vertical="center")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Coded Data"
    for i, h in enumerate(df.columns, 1):
        c = ws.cell(1, i, h); c.fill = hf; c.font = hfont; c.border = bd; c.alignment = ctr
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, h in enumerate(df.columns, 1):
            c = ws.cell(ri, ci, row[h]); c.border = bd; c.font = Font(name=AR, size=9); c.alignment = wrap
            if h == "outsider_present" and row[h] == "yes":
                c.fill = PatternFill("solid", fgColor="E2EFDA")
    widths = [12, 22, 8, 14, 16, 15, 15, 15, 10, 52, 60]
    for i, w in enumerate(widths[:len(df.columns)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # summary tab
    sm = wb.create_sheet("Summary"); n = len(df)
    present = df[df.outsider_present == "yes"]
    lines = [
        ("Interviews", n),
        ("Outsider positioning present", f"{len(present)} ({len(present)/n:.0%})"),
        ("Literal migration", f"{sum(df.literal_migration=='yes')} ({sum(df.literal_migration=='yes')/n:.0%})"),
        ("Female present", f"{sum((df.gender=='F')&(df.outsider_present=='yes'))}/{sum(df.gender=='F')}"),
        ("Male present", f"{sum((df.gender=='M')&(df.outsider_present=='yes'))}/{sum(df.gender=='M')}"),
        ("Function: resource_asset (F / M)",
            f"{sum((present.gender=='F')&(present.function=='resource_asset'))} / {sum((present.gender=='M')&(present.function=='resource_asset'))}"),
        ("Function: barrier_overcome (F / M)",
            f"{sum((present.gender=='F')&(present.function=='barrier_overcome'))} / {sum((present.gender=='M')&(present.function=='barrier_overcome'))}"),
    ]
    for r, (k, v) in enumerate(lines, 1):
        sm.cell(r, 1, k).font = Font(name=AR, bold=True, size=10)
        sm.cell(r, 2, str(v)).font = Font(name=AR, size=10)
    sm.column_dimensions["A"].width = 38; sm.column_dimensions["B"].width = 22
    wb.save(out("immigrant_outsider_coded.xlsx"))

def main():
    df = pd.read_csv(inp("immigrant_outsider_coded.csv"))
    summarise(df)
    marker_crosscheck()
    df.to_csv(out("immigrant_outsider_summary.csv"), index=False)
    build_xlsx(df)
    print("\nwrote outputs/immigrant_outsider_coded.xlsx")

if __name__ == "__main__":
    main()
